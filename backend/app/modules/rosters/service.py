import csv
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO, StringIO
from pathlib import Path
from zipfile import BadZipFile, ZipFile, is_zipfile

import xlrd
from olefile import OleFileIO
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError

from app.modules.rosters.model import Student
from app.modules.rosters.schemas import (
    ImportColumnMapping,
    RosterImportRow,
    StudentCreate,
)

MAX_IMPORT_SIZE = 10 * 1024 * 1024
MAX_UNCOMPRESSED_SIZE = 50 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
MAX_IMPORT_COLUMNS = 100
MAX_ZIP_ENTRIES = 5000

ALLOWED_MIME_TYPES: dict[str, set[str]] = {
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/octet-stream",
        "application/zip",
    },
    ".xls": {
        "application/vnd.ms-excel",
        "application/octet-stream",
        "application/x-ole-storage",
    },
    ".csv": {
        "text/csv",
        "application/csv",
        "text/plain",
        "application/vnd.ms-excel",
        "application/octet-stream",
    },
}

HEADER_ALIASES: dict[str, set[str]] = {
    "full_name": {
        "apellidos y nombres",
        "apellidos nombres",
        "apellido y nombre",
        "nombres y apellidos",
        "nombres apellidos",
        "nombre y apellidos",
        "nombre apellidos",
        "nombre completo",
        "nombres completos",
        "estudiante",
        "nombre del estudiante",
        "alumno",
        "nombre del alumno",
    },
    "first_names": {"nombres", "nombre", "primer nombre", "first names", "first name"},
    "last_names": {"apellidos", "apellido", "last names", "last name"},
    "internal_code": {
        "codigo",
        "codigo interno",
        "codigo de estudiante",
        "codigo del estudiante",
        "cod estudiante",
        "id estudiante",
    },
    "document_number": {
        "dni",
        "documento",
        "documento de identidad",
        "documento identidad",
        "numero de documento",
        "numero documento",
        "numero dni",
        "nro documento",
        "nro dni",
        "n documento",
        "num doc",
        "dni codigo",
        "dni documento",
        "dni o documento",
    },
    "sex": {"sexo", "genero", "sexo genero"},
    "notes": {
        "observacion",
        "observaciones",
        "nota",
        "notas",
        "comentario",
        "comentarios",
    },
}

EXAMPLE_MARKER_ALIASES = {
    "avendia tipo fila",
    "tipo de fila avendia",
    "fila de ejemplo no importar",
}
EXAMPLE_MARKER_VALUES = {"example", "ejemplo", "sample", "no importar"}


class ImportFileError(Exception):
    def __init__(self, message: str, *, status_code: int = 422) -> None:
        super().__init__(message)
        self.message = message
        self.status_code = status_code


@dataclass
class ParsedTable:
    file_type: str
    sheet_name: str | None
    available_sheets: list[str]
    rows: list[list[str]]


@dataclass
class PreviewData:
    columns: list[str]
    rows: list[RosterImportRow]
    suggested_mapping: ImportColumnMapping
    ignored_empty_rows: int
    ignored_example_rows: int
    warnings: list[str]


@dataclass
class PreparedImport:
    students: list[StudentCreate]
    skipped_count: int


def normalized_text(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    without_marks = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", without_marks.casefold()).strip()


def normalized_identifier(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    without_marks = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", without_marks.casefold())


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Sí" if value else "No"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value).strip()


def _validate_mime(extension: str, content_type: str | None) -> None:
    mime = (content_type or "application/octet-stream").split(";", 1)[0].strip().lower()
    if mime not in ALLOWED_MIME_TYPES[extension]:
        raise ImportFileError(
            "El tipo del archivo no coincide con su extensión. "
            "Guárdalo nuevamente como XLSX, XLS o CSV e inténtalo otra vez."
        )


def _validate_xlsx_container(content: bytes) -> None:
    if not is_zipfile(BytesIO(content)):
        raise ImportFileError("El archivo XLSX no es un libro de Excel válido o está dañado.")
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ZIP_ENTRIES:
                raise ImportFileError("El libro contiene demasiados archivos internos.")
            if sum(entry.file_size for entry in entries) > MAX_UNCOMPRESSED_SIZE:
                raise ImportFileError("El libro se expande por encima del límite seguro de 50 MB.")
            names = {entry.filename.casefold() for entry in entries}
            if "[content_types].xml" not in names or "xl/workbook.xml" not in names:
                raise ImportFileError(
                    "El archivo no contiene la estructura requerida de un libro XLSX."
                )
            macro_names = (
                "vbaproject.bin",
                "vbadata.xml",
                "macrosheet",
                "_vba_project",
            )
            if any(any(marker in name for marker in macro_names) for name in names):
                raise ImportFileError(
                    "Los libros con macros no están permitidos. "
                    "Guarda una copia como .xlsx sin macros."
                )
            content_types = archive.read("[Content_Types].xml").lower()
            if b"macroenabled" in content_types or b"vbaproject" in content_types:
                raise ImportFileError(
                    "Los libros con macros no están permitidos. "
                    "Guarda una copia como .xlsx sin macros."
                )
    except BadZipFile as exc:
        raise ImportFileError("El archivo XLSX está dañado y no puede leerse.") from exc


def _validate_xls_container(content: bytes) -> None:
    if not content.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        raise ImportFileError("El archivo XLS no coincide con el formato Excel 97-2003.")
    try:
        with OleFileIO(BytesIO(content)) as ole_file:
            stream_names = ["/".join(parts).casefold() for parts in ole_file.listdir()]
    except Exception as exc:
        raise ImportFileError("El archivo XLS está dañado y no puede leerse.") from exc
    macro_markers = ("vba", "_vba_project", "macros", "modules")
    if any(any(marker in name for marker in macro_markers) for name in stream_names):
        raise ImportFileError(
            "Los libros con macros no están permitidos. "
            "Guarda una copia sin macros o en formato .xlsx."
        )


def validate_import_file(
    filename: str | None,
    content_type: str | None,
    content: bytes,
) -> str:
    extension = Path(filename or "").suffix.casefold()
    if extension not in ALLOWED_MIME_TYPES:
        raise ImportFileError("Formato no admitido. Usa un archivo .xlsx, .xls o .csv sin macros.")
    _validate_mime(extension, content_type)
    if not content:
        raise ImportFileError("El archivo está vacío.")
    if len(content) > MAX_IMPORT_SIZE:
        raise ImportFileError(
            "El archivo supera 10 MB. Reduce la nómina o guárdala como CSV.",
            status_code=413,
        )
    if extension == ".xlsx":
        _validate_xlsx_container(content)
    elif extension == ".xls":
        _validate_xls_container(content)
    elif b"\x00" in content[:8192]:
        raise ImportFileError("El archivo CSV contiene datos binarios y no puede leerse.")
    return extension


def _bounded_rows(rows: list[list[str]]) -> list[list[str]]:
    if len(rows) > MAX_IMPORT_ROWS + 25:
        raise ImportFileError(
            f"El archivo supera el máximo de {MAX_IMPORT_ROWS:,} filas por importación."
        )
    widest = max((len(row) for row in rows), default=0)
    if widest > MAX_IMPORT_COLUMNS:
        raise ImportFileError(f"El archivo supera el máximo de {MAX_IMPORT_COLUMNS} columnas.")
    return rows


def _choose_sheet(
    requested_name: str | None,
    available_sheets: list[str],
) -> str:
    if requested_name:
        if requested_name not in available_sheets:
            raise ImportFileError("La hoja seleccionada no existe en el libro.")
        return requested_name
    return available_sheets[0]


def _parse_xlsx(content: bytes, requested_sheet: str | None) -> ParsedTable:
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        available = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        if not available:
            raise ImportFileError("El libro XLSX no tiene hojas visibles.")
        if requested_sheet:
            selected = _choose_sheet(requested_sheet, available)
        else:
            sheet_scores: list[tuple[int, int, int, str]] = []
            for sheet_index, candidate_name in enumerate(available):
                sample = [
                    [clean_cell(value) for value in row]
                    for row in workbook[candidate_name].iter_rows(
                        min_row=1,
                        max_row=25,
                        values_only=True,
                    )
                ]
                score = max((_header_score(row) for row in sample), default=0)
                nonempty = sum(any(cell for cell in row) for row in sample)
                sheet_scores.append((score, nonempty, -sheet_index, candidate_name))
            selected = max(sheet_scores)[3]
        sheet = workbook[selected]
        rows = [[clean_cell(value) for value in row] for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    except ImportFileError:
        raise
    except Exception as exc:
        raise ImportFileError(
            "No se pudo leer el XLSX. Verifica que no esté corrupto ni protegido."
        ) from exc
    return ParsedTable("xlsx", selected, available, _bounded_rows(rows))


def _parse_xls(content: bytes, requested_sheet: str | None) -> ParsedTable:
    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        available = workbook.sheet_names()
        if not available:
            raise ImportFileError("El libro XLS no contiene hojas.")
        if requested_sheet:
            selected = _choose_sheet(requested_sheet, available)
        else:
            sheet_scores = []
            for sheet_index, candidate_name in enumerate(available):
                candidate = workbook.sheet_by_name(candidate_name)
                sample = [
                    [
                        clean_cell(candidate.cell_value(row_index, column_index))
                        for column_index in range(candidate.ncols)
                    ]
                    for row_index in range(min(candidate.nrows, 25))
                ]
                score = max((_header_score(row) for row in sample), default=0)
                nonempty = sum(any(cell for cell in row) for row in sample)
                sheet_scores.append((score, nonempty, -sheet_index, candidate_name))
            selected = max(sheet_scores)[3]
        sheet = workbook.sheet_by_name(selected)
        rows = [
            [
                clean_cell(sheet.cell_value(row_index, column_index))
                for column_index in range(sheet.ncols)
            ]
            for row_index in range(sheet.nrows)
        ]
        workbook.release_resources()
    except ImportFileError:
        raise
    except (xlrd.XLRDError, IndexError, ValueError) as exc:
        raise ImportFileError(
            "No se pudo leer el XLS. Verifica que sea un archivo Excel 97-2003 real, "
            "sin contraseña ni macros."
        ) from exc
    return ParsedTable("xls", selected, available, _bounded_rows(rows))


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ImportFileError("No se pudo reconocer la codificación del CSV. Guárdalo como UTF-8.")


def _parse_csv(content: bytes) -> ParsedTable:
    decoded = _decode_csv(content)
    sample = decoded[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    try:
        reader = csv.reader(StringIO(decoded), dialect)
        rows = [[clean_cell(value) for value in row] for row in reader]
    except csv.Error as exc:
        raise ImportFileError("El CSV tiene una estructura inválida.") from exc
    return ParsedTable("csv", None, [], _bounded_rows(rows))


def parse_import_file(
    filename: str | None,
    content_type: str | None,
    content: bytes,
    *,
    sheet_name: str | None = None,
) -> ParsedTable:
    extension = validate_import_file(filename, content_type, content)
    if extension == ".xlsx":
        return _parse_xlsx(content, sheet_name)
    if extension == ".xls":
        return _parse_xls(content, sheet_name)
    return _parse_csv(content)


def _header_score(row: list[str]) -> int:
    normalized = {normalized_text(cell) for cell in row if cell.strip()}
    recognized = sum(
        1 for aliases in HEADER_ALIASES.values() if any(alias in normalized for alias in aliases)
    )
    if any(header in EXAMPLE_MARKER_ALIASES for header in normalized):
        recognized += 1
    return recognized


def _unique_headers(row: list[str]) -> list[str]:
    headers: list[str] = []
    seen: defaultdict[str, int] = defaultdict(int)
    for index, value in enumerate(row, start=1):
        base = re.sub(r"\s+", " ", value).strip() or f"Columna {index}"
        seen[base] += 1
        headers.append(base if seen[base] == 1 else f"{base} ({seen[base]})")
    while headers and normalized_text(headers[-1]).startswith("columna "):
        headers.pop()
    return headers


def _find_header_row(rows: list[list[str]]) -> int:
    candidates = [index for index, row in enumerate(rows[:25]) if any(cell.strip() for cell in row)]
    if not candidates:
        raise ImportFileError("El archivo no contiene filas con datos.")
    return max(candidates, key=lambda index: (_header_score(rows[index]), -index))


def suggest_mapping(columns: list[str]) -> ImportColumnMapping:
    mapping: dict[str, str] = {}
    for column in columns:
        normalized = normalized_text(column)
        for field, aliases in HEADER_ALIASES.items():
            if field not in mapping and normalized in aliases:
                mapping[field] = column
                break
    return ImportColumnMapping.model_validate(mapping)


def validate_mapping(mapping: ImportColumnMapping, columns: set[str]) -> None:
    selected = mapping.model_dump(exclude_none=True)
    unknown = [header for header in selected.values() if header not in columns]
    if unknown:
        raise ImportFileError(
            "Una o más columnas seleccionadas ya no están disponibles. Vuelve a revisar el mapeo."
        )
    if not mapping.full_name and not (mapping.first_names and mapping.last_names):
        raise ImportFileError(
            "Selecciona la columna de nombre completo o las columnas de nombres y apellidos."
        )


def _mapped_value(row: dict[str, str], header: str | None) -> str | None:
    if not header:
        return None
    value = re.sub(r"\s+", " ", row.get(header, "")).strip()
    return value or None


def student_from_row(row: dict[str, str], mapping: ImportColumnMapping) -> StudentCreate:
    if mapping.full_name:
        full_name = _mapped_value(row, mapping.full_name)
    else:
        last_names = _mapped_value(row, mapping.last_names)
        first_names = _mapped_value(row, mapping.first_names)
        full_name = " ".join(value for value in (last_names, first_names) if value) or None
    if not full_name:
        raise ImportFileError("Falta el nombre completo.")
    if not any(character.isalpha() for character in full_name):
        raise ImportFileError("El nombre debe contener letras.")
    try:
        return StudentCreate(
            full_name=full_name,
            internal_code=_mapped_value(row, mapping.internal_code),
            document_number=_mapped_value(row, mapping.document_number),
            sex=_mapped_value(row, mapping.sex),
            notes=_mapped_value(row, mapping.notes),
        )
    except ValidationError as exc:
        fields = {str(error["loc"][0]) for error in exc.errors() if error.get("loc")}
        labels = {
            "full_name": "nombre completo",
            "internal_code": "código interno",
            "document_number": "documento",
            "sex": "sexo",
            "notes": "observación",
        }
        invalid = ", ".join(labels.get(field, field) for field in sorted(fields))
        raise ImportFileError(f"Revisa la longitud o el formato de: {invalid}.") from exc


def _student_keys(student: StudentCreate | Student) -> set[tuple[str, str]]:
    keys = {("full_name", normalized_text(student.full_name))}
    if student.internal_code:
        keys.add(("internal_code", normalized_identifier(student.internal_code)))
    if student.document_number:
        keys.add(("document_number", normalized_identifier(student.document_number)))
    return {(field, value) for field, value in keys if value}


def _example_marker_column(columns: list[str]) -> str | None:
    return next(
        (column for column in columns if normalized_text(column) in EXAMPLE_MARKER_ALIASES),
        None,
    )


def build_preview(table: ParsedTable, existing_students: list[Student]) -> PreviewData:
    header_index = _find_header_row(table.rows)
    columns = _unique_headers(table.rows[header_index])
    if not columns:
        raise ImportFileError(
            "No se encontraron encabezados. Colócalos en una sola fila o usa la plantilla."
        )
    mapping = suggest_mapping(columns)
    marker_column = _example_marker_column(columns)
    raw_rows: list[tuple[int, dict[str, str]]] = []
    ignored_empty = 0
    ignored_example = 0
    for source_index, values in enumerate(table.rows[header_index + 1 :], start=header_index + 2):
        row = {
            column: values[index].strip() if index < len(values) else ""
            for index, column in enumerate(columns)
        }
        if not any(row.values()):
            ignored_empty += 1
            continue
        if marker_column and normalized_text(row.get(marker_column, "")) in EXAMPLE_MARKER_VALUES:
            visible_values = [value for column, value in row.items() if column != marker_column]
            if any(normalized_text(value).startswith("ejemplo") for value in visible_values):
                ignored_example += 1
                continue
        if marker_column:
            row.pop(marker_column, None)
        raw_rows.append((source_index, row))

    if len(raw_rows) > MAX_IMPORT_ROWS:
        raise ImportFileError(
            f"El archivo supera el máximo de {MAX_IMPORT_ROWS:,} estudiantes por importación."
        )

    visible_columns = [column for column in columns if column != marker_column]
    requires_mapping = not mapping.full_name and not (mapping.first_names and mapping.last_names)
    errors_by_index: dict[int, list[str]] = defaultdict(list)
    students_by_index: dict[int, StudentCreate] = {}
    if not requires_mapping:
        for index, (_, row) in enumerate(raw_rows):
            try:
                students_by_index[index] = student_from_row(row, mapping)
            except ImportFileError as exc:
                errors_by_index[index].append(exc.message)

    file_key_rows: defaultdict[tuple[str, str], list[int]] = defaultdict(list)
    for index, student in students_by_index.items():
        for key in _student_keys(student):
            file_key_rows[key].append(index)
    duplicate_file_rows = {
        index for indexes in file_key_rows.values() if len(indexes) > 1 for index in indexes
    }
    existing_keys = {key for student in existing_students for key in _student_keys(student)}
    duplicate_roster_rows = {
        index
        for index, student in students_by_index.items()
        if _student_keys(student) & existing_keys
    }

    preview_rows: list[RosterImportRow] = []
    for index, (row_number, row) in enumerate(raw_rows):
        messages = list(errors_by_index[index])
        is_duplicate = index in duplicate_file_rows or index in duplicate_roster_rows
        if index in duplicate_file_rows:
            messages.append("La fila repite un nombre, código o documento dentro del archivo.")
        if index in duplicate_roster_rows:
            messages.append("La fila coincide con un estudiante que ya existe en esta nómina.")
        if requires_mapping:
            messages.append("Selecciona manualmente las columnas que contienen los nombres.")
            status = None
        else:
            status = (
                "invalid" if errors_by_index[index] else "duplicate" if is_duplicate else "valid"
            )
        preview_rows.append(
            RosterImportRow(
                row_number=row_number,
                values=row,
                status=status,
                message=" ".join(messages) or None,
            )
        )

    warnings: list[str] = []
    if requires_mapping:
        warnings.append(
            "No se reconoció la columna de nombres. Relaciona las columnas antes de confirmar."
        )
    if ignored_empty:
        warnings.append(f"Se ignoraron {ignored_empty} filas totalmente vacías.")
    if ignored_example:
        warnings.append(f"Se ignoraron {ignored_example} filas de ejemplo de la plantilla.")
    duplicate_count = len(duplicate_file_rows | duplicate_roster_rows)
    if duplicate_count:
        warnings.append(f"Se detectaron {duplicate_count} filas duplicadas.")
    return PreviewData(
        columns=visible_columns,
        rows=preview_rows,
        suggested_mapping=mapping,
        ignored_empty_rows=ignored_empty,
        ignored_example_rows=ignored_example,
        warnings=warnings,
    )


def prepare_import(
    rows: list[dict[str, str]],
    mapping: ImportColumnMapping,
    existing_students: list[Student],
    *,
    skip_duplicates: bool,
) -> PreparedImport:
    columns = {column for row in rows for column in row}
    validate_mapping(mapping, columns)
    parsed: list[tuple[int, StudentCreate]] = []
    invalid: list[dict[str, object]] = []
    for index, row in enumerate(rows, start=1):
        if not any(value.strip() for value in row.values()):
            continue
        try:
            parsed.append((index, student_from_row(row, mapping)))
        except ImportFileError as exc:
            invalid.append({"row": index, "message": exc.message})
    if invalid:
        raise ImportFileError(
            "Hay filas inválidas. Corrige los datos indicados y vuelve a confirmar."
        )

    occupied_keys = {key for student in existing_students for key in _student_keys(student)}
    accepted: list[StudentCreate] = []
    duplicate_rows: list[int] = []
    for row_number, student in parsed:
        keys = _student_keys(student)
        if keys & occupied_keys:
            duplicate_rows.append(row_number)
            continue
        accepted.append(student)
        occupied_keys.update(keys)
    if duplicate_rows and not skip_duplicates:
        raise ImportFileError(
            "Se detectaron duplicados. Revísalos o confirma indicando que deben omitirse.",
            status_code=409,
        )
    return PreparedImport(students=accepted, skipped_count=len(duplicate_rows))


def duplicate_fields(candidate: StudentCreate, existing_students: list[Student]) -> list[str]:
    candidate_keys = _student_keys(candidate)
    conflicts = {
        field
        for student in existing_students
        for field, value in (_student_keys(student) & candidate_keys)
    }
    return sorted(conflicts)


def build_template_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Estudiantes"
    headers = [
        "N.°",
        "Apellidos y nombres",
        "Código interno",
        "DNI / documento",
        "Sexo",
        "Observación",
        "Avendia tipo fila",
    ]
    sheet.append(headers)
    sheet.append(
        [
            "1",
            "Ejemplo: Quispe Flores, Ana María",
            "EST-001",
            "12345678",
            "F",
            "Fila de ejemplo; reemplázala con un estudiante real.",
            "example",
        ]
    )
    header_fill = PatternFill("solid", fgColor="2563EB")
    example_fill = PatternFill("solid", fgColor="EDE9FE")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for cell in sheet[2]:
        cell.fill = example_fill
    sheet["B2"].comment = Comment(
        "Esta fila es solo una guía y Avendia no la importará. Reemplázala o agrega filas debajo.",
        "Avendia",
    )
    widths = {"A": 8, "B": 42, "C": 20, "D": 20, "E": 12, "F": 48, "G": 22}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.column_dimensions["G"].hidden = True
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:F5001"
    sex_validation = DataValidation(type="list", formula1='"F,M,Otro,Prefiero no indicar"')
    sex_validation.error = "Selecciona una opción de la lista."
    sex_validation.errorTitle = "Valor no válido"
    sheet.add_data_validation(sex_validation)
    sex_validation.add("E3:E5001")

    instructions = workbook.create_sheet("Instrucciones")
    instructions.append(["Plantilla de nómina Avendia"])
    instructions.append(["Una fila corresponde a un estudiante."])
    instructions.append(["Apellidos y nombres es el único campo obligatorio."])
    instructions.append(["No incluyas diagnósticos médicos ni información sensible."])
    instructions.append(["La fila violeta es un ejemplo y no se importará."])
    instructions["A1"].font = Font(size=16, bold=True, color="4F46E5")
    instructions.column_dimensions["A"].width = 75
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
