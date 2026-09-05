from io import BytesIO
from uuid import uuid4
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
import xlwt
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook

from app.main import app


def registration(email: str) -> dict[str, object]:
    return {
        "email": email,
        "full_name": "María Gómez",
        "password": "a-secure-password",
        "dre": "DRE Lima Metropolitana",
        "ugel": "UGEL 03",
        "school_name": "I.E. José María Arguedas",
        "director_name": "Elena Torres",
        "education_modality": "EBR",
        "education_level": "Secundaria",
        "grade": "3° de Secundaria",
        "section": "A",
        "curricular_area": "Comunicación",
        "school_year": 2026,
    }


async def authenticated_headers(client: AsyncClient, email: str) -> dict[str, str]:
    payload = registration(email)
    assert (await client.post("/api/v1/auth/register", json=payload)).status_code == 201
    login = await client.post(
        "/api/v1/auth/login",
        json={"email": email, "password": payload["password"]},
    )
    assert login.status_code == 200
    return {"Authorization": f"Bearer {login.json()['access_token']}"}


def roster_payload(**overrides: object) -> dict[str, object]:
    payload: dict[str, object] = {
        "school_year": 2026,
        "institution_name": "I.E. José María Arguedas",
        "modality": "EBR",
        "education_level": "Secundaria",
        "grade": "3° de Secundaria",
        "section": "A",
        "name": "Tercero A",
    }
    payload.update(overrides)
    return payload


async def create_roster(client: AsyncClient, headers: dict[str, str]) -> dict[str, object]:
    response = await client.post("/api/v1/rosters", headers=headers, json=roster_payload())
    assert response.status_code == 201, response.text
    return response.json()


def xlsx_file(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Nómina"
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def xlsx_file_with_instructions(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instrucciones"
    instructions.append(["Completa la hoja Nómina con una fila por estudiante."])
    sheet = workbook.create_sheet("Nómina")
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def xls_file(rows: list[list[object]]) -> bytes:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Nómina")
    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            sheet.write(row_index, column_index, value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.mark.asyncio
async def test_roster_crud_soft_delete_reorder_and_owner_isolation() -> None:
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        owner = await authenticated_headers(client, "roster-owner@example.edu")
        other = await authenticated_headers(client, "roster-other@example.edu")
        roster = await create_roster(client, owner)
        roster_id = roster["id"]

        listing = await client.get("/api/v1/rosters", headers=owner)
        assert listing.status_code == 200
        assert listing.json()["total"] == 1
        assert listing.json()["items"][0]["active_student_count"] == 0

        for path in (
            f"/api/v1/rosters/{roster_id}",
            f"/api/v1/rosters/{roster_id}/students",
        ):
            assert (await client.get(path, headers=other)).status_code == 404
        forbidden_update = await client.patch(
            f"/api/v1/rosters/{roster_id}",
            headers=other,
            json={"name": "No permitido"},
        )
        assert forbidden_update.status_code == 404

        first = await client.post(
            f"/api/v1/rosters/{roster_id}/students",
            headers=owner,
            json={
                "full_name": "Quispe Flores, Ana María",
                "internal_code": "EST-001",
                "document_number": "12345678",
                "notes": "Participa activamente.",
            },
        )
        second = await client.post(
            f"/api/v1/rosters/{roster_id}/students",
            headers=owner,
            json={"full_name": "Rojas León, Luis", "internal_code": "EST-002"},
        )
        assert first.status_code == second.status_code == 201

        duplicate = await client.post(
            f"/api/v1/rosters/{roster_id}/students",
            headers=owner,
            json={"full_name": "  QUÍSPE   FLORES, ANA MARÍA  "},
        )
        assert duplicate.status_code == 409
        assert "full_name" in duplicate.json()["detail"]["duplicate_fields"]

        second_id = second.json()["id"]
        first_id = first.json()["id"]
        reordered = await client.post(
            f"/api/v1/rosters/{roster_id}/students/reorder",
            headers=owner,
            json={"student_ids": [second_id, first_id]},
        )
        assert reordered.status_code == 200
        assert [item["id"] for item in reordered.json()["items"]] == [second_id, first_id]
        assert [item["sort_order"] for item in reordered.json()["items"]] == [0, 1]

        foreign_reorder = await client.post(
            f"/api/v1/rosters/{roster_id}/students/reorder",
            headers=owner,
            json={"student_ids": [second_id, str(uuid4())]},
        )
        assert foreign_reorder.status_code == 422

        assert (
            await client.delete(
                f"/api/v1/rosters/{roster_id}/students/{first_id}",
                headers=owner,
            )
        ).status_code == 204
        active_students = await client.get(
            f"/api/v1/rosters/{roster_id}/students",
            headers=owner,
        )
        inactive_students = await client.get(
            f"/api/v1/rosters/{roster_id}/students?active=false",
            headers=owner,
        )
        assert [item["id"] for item in active_students.json()["items"]] == [second_id]
        assert [item["id"] for item in inactive_students.json()["items"]] == [first_id]

        assert (
            await client.delete(f"/api/v1/rosters/{roster_id}", headers=owner)
        ).status_code == 204
        assert (await client.get("/api/v1/rosters", headers=owner)).json()["total"] == 0
        inactive_rosters = await client.get("/api/v1/rosters?active=false", headers=owner)
        assert inactive_rosters.json()["items"][0]["id"] == roster_id
        blocked_add = await client.post(
            f"/api/v1/rosters/{roster_id}/students",
            headers=owner,
            json={"full_name": "Estudiante bloqueado"},
        )
        assert blocked_add.status_code == 409


@pytest.mark.asyncio
async def test_roster_education_selection_is_validated_on_create_and_patch() -> None:
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        headers = await authenticated_headers(client, "roster-catalog@example.edu")
        invalid = await client.post(
            "/api/v1/rosters",
            headers=headers,
            json=roster_payload(education_level="Primaria", grade="3° de Secundaria"),
        )
        assert invalid.status_code == 422

        roster = await create_roster(client, headers)
        invalid_patch = await client.patch(
            f"/api/v1/rosters/{roster['id']}",
            headers=headers,
            json={"modality": "EBA"},
        )
        assert invalid_patch.status_code == 422


@pytest.mark.asyncio
async def test_csv_preview_confirm_duplicates_and_atomic_validation() -> None:
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        headers = await authenticated_headers(client, "roster-csv@example.edu")
        roster = await create_roster(client, headers)
        roster_id = roster["id"]
        content = (
            "Apellidos y nombres;DNI;Observación\n"
            "Quispe Flores, Ana María;12345678;Participa\n"
            ";;\n"
            "Rojas León, Luis;87654321;\n"
        ).encode()
        preview = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/preview",
            headers=headers,
            files={"file": ("estudiantes.csv", content, "text/csv")},
        )
        assert preview.status_code == 200, preview.text
        body = preview.json()
        assert body["suggested_mapping"] == {
            "full_name": "Apellidos y nombres",
            "first_names": None,
            "last_names": None,
            "internal_code": None,
            "document_number": "DNI",
            "sex": None,
            "notes": "Observación",
        }
        assert body["total_rows"] == 2
        assert body["ignored_empty_rows"] == 1
        assert all("values" in row for row in body["rows"])
        assert all(row["status"] == "valid" for row in body["rows"])

        confirmation = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/confirm",
            headers=headers,
            json={
                "mapping": body["suggested_mapping"],
                "rows": [row["values"] for row in body["rows"]],
                "skip_duplicates": True,
            },
        )
        assert confirmation.status_code == 201, confirmation.text
        assert confirmation.json()["created_count"] == 2
        assert confirmation.json()["skipped_count"] == 0
        assert len(confirmation.json()["students"]) == 2

        formatted_document_duplicate = await client.post(
            f"/api/v1/rosters/{roster_id}/students",
            headers=headers,
            json={
                "full_name": "Otro Nombre",
                "document_number": "12.345.678",
            },
        )
        assert formatted_document_duplicate.status_code == 409
        assert (
            "document_number" in formatted_document_duplicate.json()["detail"]["duplicate_fields"]
        )

        repeated = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/preview",
            headers=headers,
            files={"file": ("estudiantes.csv", content, "text/csv")},
        )
        assert {row["status"] for row in repeated.json()["rows"]} == {"duplicate"}
        skipped = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/confirm",
            headers=headers,
            json={
                "mapping": repeated.json()["suggested_mapping"],
                "rows": [row["values"] for row in repeated.json()["rows"]],
                "skip_duplicates": True,
            },
        )
        assert skipped.status_code == 201
        assert skipped.json()["created_count"] == 0
        assert skipped.json()["skipped_count"] == 2

        duplicate_file = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/preview",
            headers=headers,
            files={
                "file": (
                    "duplicados.csv",
                    (
                        b"Nombre completo,DNI\n"
                        b"Nueva Estudiante,11223344\n"
                        b"Nueva Estudiante,11223344\n"
                    ),
                    "text/csv",
                )
            },
        )
        assert duplicate_file.status_code == 200
        assert [row["status"] for row in duplicate_file.json()["rows"]] == [
            "duplicate",
            "duplicate",
        ]

        unusual_header = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/preview",
            headers=headers,
            files={"file": ("manual.csv", b"Dato\nPersona Manual\n", "text/csv")},
        )
        assert unusual_header.status_code == 200
        assert unusual_header.json()["suggested_mapping"]["full_name"] is None
        assert unusual_header.json()["rows"][0]["status"] is None
        manual_confirm = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/confirm",
            headers=headers,
            json={
                "mapping": {"full_name": "Dato"},
                "rows": [unusual_header.json()["rows"][0]["values"]],
                "skip_duplicates": True,
            },
        )
        assert manual_confirm.status_code == 201
        assert manual_confirm.json()["created_count"] == 1

        atomic_failure = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/confirm",
            headers=headers,
            json={
                "mapping": {"full_name": "Nombre"},
                "rows": [
                    {"Nombre": "Estudiante válido", "Otro": ""},
                    {"Nombre": "", "Otro": "dato sin nombre"},
                ],
                "skip_duplicates": False,
            },
        )
        assert atomic_failure.status_code == 422
        final_list = await client.get(
            f"/api/v1/rosters/{roster_id}/students",
            headers=headers,
        )
        assert final_list.json()["total"] == 3


@pytest.mark.asyncio
async def test_real_xlsx_and_xls_preview_and_template_round_trip() -> None:
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        headers = await authenticated_headers(client, "roster-excel@example.edu")
        roster = await create_roster(client, headers)
        roster_id = roster["id"]
        cases = [
            (
                "nomina.xlsx",
                xlsx_file_with_instructions(
                    [
                        ["N.°", "Apellidos", "Nombres", "Código", "Sexo"],
                        [1, "Mamani Soto", "Elena", "A-001", "F"],
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "xlsx",
                "A-001",
            ),
            (
                "nomina.xls",
                xls_file(
                    [
                        ["N.°", "Apellidos", "Nombres", "Código", "Sexo"],
                        [1, "Paredes Luna", "José", "A-002", "M"],
                    ]
                ),
                "application/vnd.ms-excel",
                "xls",
                "A-002",
            ),
        ]
        for filename, content, content_type, expected_type, expected_code in cases:
            preview = await client.post(
                f"/api/v1/rosters/{roster_id}/imports/preview",
                headers=headers,
                files={"file": (filename, content, content_type)},
            )
            assert preview.status_code == 200, preview.text
            assert preview.json()["file_type"] == expected_type
            assert preview.json()["sheet_name"] == "Nómina"
            assert preview.json()["suggested_mapping"]["last_names"] == "Apellidos"
            assert preview.json()["suggested_mapping"]["first_names"] == "Nombres"
            assert preview.json()["rows"][0]["values"]["Código"] == expected_code
            confirmation = await client.post(
                f"/api/v1/rosters/{roster_id}/imports/confirm",
                headers=headers,
                json={
                    "mapping": preview.json()["suggested_mapping"],
                    "rows": [row["values"] for row in preview.json()["rows"]],
                    "skip_duplicates": True,
                },
            )
            assert confirmation.status_code == 201, confirmation.text
            assert confirmation.json()["created_count"] == 1

        template = await client.get("/api/v1/rosters/template", headers=headers)
        assert template.status_code == 200
        workbook = load_workbook(BytesIO(template.content), read_only=True)
        assert workbook.sheetnames == ["Estudiantes", "Instrucciones"]
        assert workbook["Estudiantes"]["B2"].value.startswith("Ejemplo:")
        workbook.close()
        template_preview = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/preview",
            headers=headers,
            files={"file": ("plantilla.xlsx", template.content, template.headers["content-type"])},
        )
        assert template_preview.status_code == 200, template_preview.text
        assert template_preview.json()["total_rows"] == 0
        assert template_preview.json()["ignored_example_rows"] == 1
        assert template_preview.json()["suggested_mapping"]["document_number"] == "DNI / documento"


@pytest.mark.asyncio
async def test_import_rejects_macros_mime_mismatch_corruption_and_foreign_roster() -> None:
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        owner = await authenticated_headers(client, "roster-files-owner@example.edu")
        other = await authenticated_headers(client, "roster-files-other@example.edu")
        roster = await create_roster(client, owner)
        roster_id = roster["id"]

        macro_book = BytesIO(xlsx_file([["Nombre completo"], ["Ana Torres"]]))
        with ZipFile(macro_book, mode="a", compression=ZIP_DEFLATED) as archive:
            archive.writestr("xl/vbaProject.bin", b"not-a-real-macro")
        macro = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/preview",
            headers=owner,
            files={
                "file": (
                    "nomina.xlsx",
                    macro_book.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert macro.status_code == 422
        assert "macros" in macro.text.lower()

        mismatch = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/preview",
            headers=owner,
            files={"file": ("nomina.xlsx", b"Nombre\nAna", "text/csv")},
        )
        assert mismatch.status_code == 422

        corrupt_xls = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/preview",
            headers=owner,
            files={"file": ("nomina.xls", b"not-an-ole-file", "application/vnd.ms-excel")},
        )
        assert corrupt_xls.status_code == 422

        oversized = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/preview",
            headers=owner,
            files={"file": ("nomina.csv", b"x" * (10 * 1024 * 1024 + 1), "text/csv")},
        )
        assert oversized.status_code == 413

        foreign = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/preview",
            headers=other,
            files={"file": ("nomina.csv", b"Nombre completo\nAna", "text/csv")},
        )
        assert foreign.status_code == 404
        foreign_confirm = await client.post(
            f"/api/v1/rosters/{roster_id}/imports/confirm",
            headers=other,
            json={
                "mapping": {"full_name": "Nombre completo"},
                "rows": [{"Nombre completo": "Ana Torres"}],
                "skip_duplicates": True,
            },
        )
        assert foreign_confirm.status_code == 404

        students = await client.get(
            f"/api/v1/rosters/{roster_id}/students",
            headers=owner,
        )
        assert students.json()["total"] == 0
